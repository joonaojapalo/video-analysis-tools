from openpyxl import load_workbook

from shellcolors import print_warn

def validate_xlsx(path):
    try:
        wb = load_workbook(path)
    except:
        return ["Cannot open: %s" % path]

    if len(wb.worksheets) > 1:
        try:
            wb.sheetnames.index("Sync")
            return None
        except:
            return ["Multiple sheets found and none has name 'Sync': %s" % path]
    return None

def read_index_xlsx(path, cols):
    if not cols:
        print_warn("No columns defined to read from Excel sheet")

    # load excel ...
    wb = load_workbook(path, data_only=True)

    # check sheets
    sheet_idx = 0
    if len(wb.worksheets) > 1:
        sheet_idx = wb.sheetnames.index("Sync")
    sheet = wb.worksheets[sheet_idx]
    headers = dict((header, i + 1)
                   for i, header in enumerate(next(sheet.values)) if header is not None)

    # validate required column headers
    missing = set(cols).difference(set(headers.keys()))
    if missing:
        missing_str = ",".join(missing)
        raise Exception(
            "Index file '%s' sheet '%s' missing required header(s): %s" % (path, sheet.title, missing_str))

    row_num = 2
    rows = []
    empty_rows = 0

    while True:
        values = [sheet.cell(row_num, headers[col]).value for col in cols]

        if all(False if v is None else True for v in values):
            rows.append(values)
            empty_rows = 0
        else:
            empty_rows += 1
        row_num += 1
        if empty_rows > 2:
            break

    return rows, cols
